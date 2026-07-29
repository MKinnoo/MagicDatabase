from typing import Any

from openpyxl import load_workbook

from config import BASE_REFERENCE, SHEET_NAME
from models.card import Card
from models.translation import Translation


class ExcelRepository:

    """
    Repository chargé de manipuler Base_Reference.xlsx.
    """

    def __init__(self) -> None:

        self.workbook = load_workbook(BASE_REFERENCE)

        self.sheet = self.workbook[SHEET_NAME]

        self.headers = self.__load_headers()

    def __load_headers(self) -> dict[str, int]:

        headers: dict[str, int] = {}

        for cell in self.sheet[1]:
            headers[cell.value] = cell.column

        return headers
    
    def __get(self, row: int, column_name: str) -> Any:

        column = self.headers[column_name]

        return self.sheet.cell(row=row, column=column).value

    def __set(
        self,
        row: int,
        column_name: str,
        value: Any
    ) -> None:

        column = self.headers[column_name]

        self.sheet.cell(
            row=row,
            column=column
        ).value = value

    def find_all(self) -> list[Card]:

        """
        Charge toutes les cartes présentes dans le fichier Excel.

        Returns:
            La liste des cartes trouvées dans la feuille de calcul.
        """

        cards: list[Card] = []

        for row in range(2, self.sheet.max_row + 1):

            card = Card(

                row=row,

                card_id=self.__get(row, "CardID"),

                set_code=self.__get(row, "Set"),

                collector_number=self.__normalize_collector_number(
                    self.__get(row, "CollectorNo")
                ),

                name_en=self.__get(row, "Name_EN"),

                name_fr=self.__get(row, "Name_FR"),

                oracle_text_en=self.__get(row, "OracleText_EN"),

                oracle_text_fr=self.__get(row, "OracleText_FR")

            )

            cards.append(card)

        return cards
    
    def update_translation(

        self,

        card: Card,

        translation: Translation

    ) -> None:

        if translation.name:

            self.__set(
                card.row,
                "Name_FR",
                translation.name
            )

        if translation.oracle_text:

            self.__set(
                card.row,
                "OracleText_FR",
                translation.oracle_text
            )

    def save(self) -> None:

        self.workbook.save(BASE_REFERENCE)

    def close(self) -> None:

        self.workbook.close()

    def __normalize_collector_number(self, value: Any) -> str:
        """
        Normalise un numéro de collection provenant d'Excel.

        Exemples :
            27      -> "27"
            27.0    -> "27"
            27.5    -> "27.5"
            "27a"   -> "27a"
            "005"   -> "005"
            None    -> ""
        """

        if value is None:
            return ""

        # Si Excel l'a lu comme un entier
        if isinstance(value, int):
            return str(value)

        # Si Excel l'a lu comme un flottant
        if isinstance(value, float):

            # 27.0 devient "27"
            if value.is_integer():
                return str(int(value))

            # Cas très rare, mais on le gère proprement
            return str(value)

        # Chaîne de caractères ou autre
        return str(value).strip()
