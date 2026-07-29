from openpyxl import load_workbook


class ExcelService:

    def __init__(self, filename):

        self.workbook = load_workbook(filename)
        self.sheet = self.workbook["Cards"]

        # Création d'un dictionnaire des colonnes
        self.headers = {}

        for cell in self.sheet[1]:
            self.headers[cell.value] = cell.column

    def get_row_count(self):
        return self.sheet.max_row

    def get_value(self, row, column_name):
        column = self.headers[column_name]
        return self.sheet.cell(row=row, column=column).value

    def set_value(self, row, column_name, value):
        column = self.headers[column_name]
        self.sheet.cell(row=row, column=column).value = value

    def save(self):
        self.workbook.save(self.workbook.filename)